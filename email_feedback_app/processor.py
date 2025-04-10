import openpyxl
import os
from openpyxl.utils import column_index_from_string


def column_letter_to_index(letter):
    return column_index_from_string(letter)

def get_cell_value(sheet, row, column):
    return sheet.cell(row=row, column=column).value

def process_feedbacks(filepath, config):
    account_name = get_account_name_from_filename(filepath)
    account_config = config['accounts'].get(account_name)

    if not account_config:
        print(f"No config found for account: {account_name}")
        return

    sheet_name = account_config.get("sheet_name", "Sheet1")
    header_row = account_config.get("header_row", 1)

    wb = openpyxl.load_workbook(filepath, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"[ERROR] The sheet '{sheet_name}' was not found in the file {filepath}.")
        return

    sheet = wb[sheet_name]

    headers = {cell.value: col_idx for col_idx, cell in enumerate(sheet[header_row], start=1)}

    feedbacks = []

    for row in sheet.iter_rows(min_row=header_row + 1):
        values = {}
        row_num = row[0].row 

        if 'assignment_group' in account_config:
            group_col = column_letter_to_index(account_config['assignment_group']['column'])
            group_val = get_cell_value(sheet, row_num, group_col)
            required = account_config['assignment_group']['required_value']
            if isinstance(required, list):
                if group_val not in required:
                    continue
            else:
                if group_val != required:
                    continue

        if 'rating_text' in account_config:
            rating_col = column_letter_to_index(account_config['rating_text']['column'])
            rating_val = get_cell_value(sheet, row_num, rating_col)
            if rating_val != account_config['rating_text']['positive_value']:
                continue

        elif 'rating_inverted' in account_config:
            rating_col = column_letter_to_index(account_config['rating_inverted']['column'])
            rating_val = get_cell_value(sheet, row_num, rating_col)
            if rating_val not in account_config['rating_inverted']['valid_values']:
                continue

        else:
            rating_col = column_letter_to_index(account_config['rating'])
            rating_val = get_cell_value(sheet, row_num, rating_col)
            try:
                if float(rating_val) < 4:
                    continue
            except (ValueError, TypeError):
                continue

        values["ticket_id"] = get_cell_value(sheet, row_num, column_letter_to_index(account_config["ticket_id"]))
        values["message"] = get_cell_value(sheet, row_num, column_letter_to_index(account_config["message"]))
        values["analyst_name"] = get_cell_value(sheet, row_num, column_letter_to_index(account_config["analyst_name"]))

        if "user_name_parts" in account_config:
            parts = [get_cell_value(sheet, row_num, column_letter_to_index(col)) for col in account_config["user_name_parts"]]
            values["user_name"] = " ".join(filter(None, parts))
        elif "user_name" in account_config:
            values["user_name"] = get_cell_value(sheet, row_num, column_letter_to_index(account_config["user_name"]))
        else:
            values["user_name"] = None

        feedbacks.append(values)

    return feedbacks

def get_account_name_from_filename(filename):
    base = os.path.basename(filename)
    name, _ = os.path.splitext(base)
    return name
